from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.main import app

client = TestClient(app)


def test_download_batch_output_ok():
    proc = client.post("/api/process-batch", json={"root_folder": "dl-test"})
    assert proc.status_code == 202
    job_id = proc.json()["job_id"]

    status_r = client.get(f"/api/process-batch/status/{job_id}")
    assert status_r.status_code == 200
    st = status_r.json()
    assert st["status"] == "completed"
    download_url = st["download_url"]
    assert download_url.startswith("/api/download/batch-output/")
    basename = Path(st["output_file"]).name
    assert download_url.endswith(basename)

    head = client.head(download_url)
    assert head.status_code == 200
    assert head.headers.get("content-length")

    resp = client.get(download_url)
    assert resp.status_code == 200
    assert resp.headers.get("content-type", "").startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    wb = load_workbook(Path(st["output_file"]))
    assert "ParsedRows" in wb.sheetnames
    assert "attachment" in resp.headers.get("content-disposition", "").lower()


def test_download_batch_output_not_found():
    resp = client.get("/api/download/batch-output/does-not-exist-999999.xlsx")
    assert resp.status_code == 404
    assert resp.json()["detail"]


def test_download_batch_output_path_traversal_rejected():
    resp = client.get("/api/download/batch-output/../.env")
    assert resp.status_code == 404
