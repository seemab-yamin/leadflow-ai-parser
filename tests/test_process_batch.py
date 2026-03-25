from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.main import app

client = TestClient(app)


def _wait_completed(job_id: str) -> dict:
    """TestClient runs background tasks after POST; usually one GET is enough."""
    for _ in range(50):
        r = client.get(f"/api/process-batch/status/{job_id}")
        assert r.status_code == 200
        data = r.json()
        if data["status"] in ("completed", "failed"):
            return data
    raise AssertionError("job did not finish")


def test_process_batch_accepts_and_completes():
    resp = client.post(
        "/api/process-batch",
        json={"root_folder": "my-folder"},
    )
    assert resp.status_code == 202
    assert resp.headers.get("content-type", "").startswith("application/json")
    body = resp.json()
    assert body.get("status") == "queued"
    assert body.get("job_id")
    assert "/api/process-batch/status/" in body.get("status_url", "")

    done = _wait_completed(body["job_id"])
    assert done["status"] == "completed"
    assert isinstance(done.get("failed_paths"), list)
    assert done.get("pdf_paths_attempted") == 0
    assert done.get("pdf_paths_failed") == 0
    assert done.get("download_url")
    assert done["download_url"].startswith("/api/download/batch-output/")
    out_path = Path(done["output_file"])
    assert out_path.is_file()
    assert out_path.name.endswith(".xlsx")


def test_process_batch_status_unknown_job_404():
    resp = client.get("/api/process-batch/status/no-such-job-id-xxxxxxxx")
    assert resp.status_code == 404


def test_process_batch_missing_root_422():
    resp = client.post("/api/process-batch", json={})
    assert resp.status_code == 422


def test_process_batch_with_pdf_relative_paths_in_output():
    """DC parser stub returns no rows; output workbook is still written."""
    resp = client.post(
        "/api/process-batch",
        json={
            "root_folder": "my-batch",
            "pdf_paths": ["DC/a.pdf", "DC/sub/b.pdf", "Other/ignored.pdf"],
        },
    )
    assert resp.status_code == 202
    done = _wait_completed(resp.json()["job_id"])
    assert done["status"] == "completed"
    assert isinstance(done.get("failed_paths"), list)
    out_path = Path(done["output_file"])
    assert out_path.is_file()
    wb = load_workbook(out_path)
    parsed = wb["ParsedRows"]
    headers = list(next(parsed.iter_rows(min_row=1, max_row=1, values_only=True)))
    assert "info" in headers


def test_process_batch_400_when_no_supported_category_paths():
    resp = client.post(
        "/api/process-batch",
        json={
            "root_folder": "my-batch",
            "pdf_paths": ["a.pdf", "Other/b.pdf"],
        },
    )
    assert resp.status_code == 400
    assert "DC" in resp.json().get("detail", "")


def test_process_batch_rejects_parent_segment_in_pdf_paths_422():
    resp = client.post(
        "/api/process-batch",
        json={"root_folder": "x", "pdf_paths": ["../evil.pdf"]},
    )
    assert resp.status_code == 422
